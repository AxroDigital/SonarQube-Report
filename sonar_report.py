import requests
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

SONARQUBE_URL = os.getenv('SONARQUBE_URL', 'http://your-sonarqube-server:9000')
SONARQUBE_TOKEN = os.getenv('SONARQUBE_TOKEN')
APP_NAME = os.getenv('APP_NAME', 'TEST PROJECT')

YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
LIGHT_RED_FILL = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
LIGHT_BLUE_FILL = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
GREEN_FILL = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

def fetch_projects():
    url = f"{SONARQUBE_URL}/api/projects/search"
    response = requests.get(url, auth=(SONARQUBE_TOKEN, ''))
    
    if response.status_code == 200:
        try:
            projects = response.json().get('components', [])
            print(f"Fetched {len(projects)} projects.")
            return projects
        except requests.exceptions.JSONDecodeError as e:
            print(f"JSON decode error: {e}")
            return []
    else:
        print(f"Failed to fetch projects: {response.status_code}")
        return []

def fetch_issue_count(project_key, severity):
    url = f"{SONARQUBE_URL}/api/issues/search"
    params = {
        'componentKeys': project_key,
        'severities': severity,
        'resolved': 'false'
    }
    
    print(f"Requesting issues for project '{project_key}' with severity '{severity}'")
    response = requests.get(url, auth=(SONARQUBE_TOKEN, ''), params=params)

    if response.status_code == 200:
        issues = response.json()
        return issues.get('total', 0)
    else:
        print(f"Failed to fetch details for project {project_key}: {response.status_code}, Response: {response.text}")
        return 0

def fetch_project_metrics(project_key):
    url = f"{SONARQUBE_URL}/api/measures/component"
    params = {
        'component': project_key,
        'metricKeys': 'coverage,ncloc,bugs,vulnerabilities,code_smells,duplicated_lines_density'
    }
    
    print(f"Fetching metrics for project: {project_key}")
    response = requests.get(url, auth=(SONARQUBE_TOKEN, ''), params=params)

    if response.status_code == 200:
        metrics = response.json().get('component', {}).get('measures', [])
        metrics_dict = {metric['metric']: metric['value'] for metric in metrics}
        return metrics_dict
    else:
        print(f"Failed to fetch metrics for project {project_key}: {response.status_code}, Response: {response.text}")
        return {}

def create_excel_report(projects):
    report_data = []

    for project in projects:
        project_key = project['key']
        project_name = project['name']
        
        blocker_count = fetch_issue_count(project_key, 'BLOCKER')
        critical_count = fetch_issue_count(project_key, 'CRITICAL')
        major_count = fetch_issue_count(project_key, 'MAJOR')
        minor_count = fetch_issue_count(project_key, 'MINOR')

        metrics = fetch_project_metrics(project_key)
        coverage = metrics.get('coverage', 'N/A')
        bugs = int(metrics.get('bugs', 0))
        vulnerabilities = int(metrics.get('vulnerabilities', 0))
        duplications = metrics.get('duplicated_lines_density', 'N/A')

        report_data.append({
            'Project Key': project_key,
            'Project Name': project_name,
            'Blocker Issues': blocker_count,
            'Critical Issues': critical_count,
            'Major Issues': major_count,
            'Minor Issues': minor_count,
            'Code Coverage %': coverage,
            'Bugs': bugs,
            'Vulnerabilities': vulnerabilities,
            'Duplications %': duplications
        })

    if report_data:
        df = pd.DataFrame(report_data)
        timestamp = pd.Timestamp.now().strftime('%Y-%m-%d')

        app_name_for_filename = APP_NAME.replace(" ", "_")
        report_filename = f'/opt/sonar/reports/{app_name_for_filename}_SonarQube_Report_{timestamp.replace("-", "")}.xlsx'
        
        with pd.ExcelWriter(report_filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sonar Report', startrow=2)

            workbook = writer.book
            worksheet = writer.sheets['Sonar Report']

            title = f'{APP_NAME} - SonarQube Code Quality Report - {timestamp}'
            title_cell = worksheet.cell(row=1, column=1, value=title)
            title_cell.font = Font(size=16, bold=True, color='800080')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

            thin = Side(border_style='thin', color='000000')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            headers = ['Project Key', 'Project Name', 'Blocker Issues', 'Critical Issues', 'Major Issues', 'Minor Issues', 'Code Coverage %', 'Bugs', 'Vulnerabilities', 'Duplications %']
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if header in ['Project Key', 'Project Name', 'Code Coverage %', 'Bugs', 'Vulnerabilities', 'Duplications %']:
                    cell.fill = YELLOW_FILL
                else:
                    cell.fill = LIGHT_RED_FILL if header == 'Blocker Issues' else (
                        LIGHT_RED_FILL if header == 'Critical Issues' else (
                            LIGHT_BLUE_FILL if header == 'Major Issues' else (
                                LIGHT_BLUE_FILL if header == 'Minor Issues' else YELLOW_FILL
                            )
                        )
                    )
                cell.border = border

            for row in worksheet.iter_rows(min_row=4, max_row=len(report_data)+3, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if cell.column == 3:
                        cell.fill = GREEN_FILL if cell.value == 0 else LIGHT_RED_FILL
                    elif cell.column == 4:
                        cell.fill = GREEN_FILL if cell.value == 0 else LIGHT_RED_FILL
                    elif cell.column == 5:
                        cell.fill = GREEN_FILL if cell.value == 0 else LIGHT_BLUE_FILL
                    elif cell.column == 6:
                        cell.fill = GREEN_FILL if cell.value == 0 else LIGHT_BLUE_FILL
                    elif cell.column == 7:
                        try:
                            coverage = float(cell.value)
                            cell.fill = GREEN_FILL if coverage >= 80.0 else LIGHT_RED_FILL
                        except ValueError:
                            cell.fill = LIGHT_RED_FILL
                    elif cell.column in [8, 9]:
                        try:
                            value = int(cell.value)
                            cell.fill = GREEN_FILL if value == 0 else LIGHT_BLUE_FILL
                        except ValueError:
                            cell.fill = LIGHT_BLUE_FILL
                    elif cell.column == 10:
                        try:
                            duplications = float(cell.value)
                            cell.fill = GREEN_FILL if duplications == 0 else LIGHT_BLUE_FILL
                        except ValueError:
                            cell.fill = LIGHT_BLUE_FILL

            hyperlink_cell = worksheet.cell(row=len(report_data) + 4, column=1)
            hyperlink_cell.value = "Click here to view more"
            hyperlink_cell.hyperlink = f"{SONARQUBE_URL}"
            hyperlink_cell.font = Font(color="0000FF", underline="single")
            hyperlink_cell.alignment = Alignment(horizontal='center')

            max_project_key_length = max(len(str(row['Project Key'])) for row in report_data)
            worksheet.column_dimensions[get_column_letter(1)].width = max_project_key_length + 2
            for column_cells in worksheet.columns:
                if column_cells[0].column != 1:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        print(f"Report generated: {report_filename}")
    else:
        print("No project data available for report.")

if __name__ == "__main__":
    projects = fetch_projects()
    create_excel_report(projects)
